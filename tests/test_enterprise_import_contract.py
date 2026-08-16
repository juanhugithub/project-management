#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""企业 Excel 批量导入的模板、受控确认与前端接入契约。"""

import base64
import io
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import app
from conftest import db_conn
from enterprise_excel import ENTERPRISE_HEADERS, build_template, normalized_rows
from imports.controlled import ImportWorkflow


def _workflow(tmp_db, tmp_path):
    """使用临时数据库和归档目录，确保测试不接触正式台账。"""
    return ImportWorkflow(str(tmp_db), tmp_path / "enterprise-imports")


def _enterprise(code="91320583ENTERPRISE01", name="批量导入企业"):
    """构造覆盖新增企业表单全部字段的标准化企业行。"""
    return {
        "name": name,
        "credit_code": code,
        "enterprise_type": "高新技术企业",
        "district": "开发区",
        "qualifications": "专精特新",
        "contact_person": "张三",
        "contact_phone": "13800000000",
        "address": "测试地址",
        "note": "批量导入",
    }


def test_enterprise_template_matches_form_fields(tmp_path):
    """企业模板必须完整覆盖手动表单字段，并能解析回标准结构。"""
    target = tmp_path / "企业导入模板.xlsx"
    build_template(target)
    workbook = load_workbook(target)
    sheet = workbook["企业台账"]
    assert [cell.value.removesuffix("*") for cell in sheet[1]] == [item[0] for item in ENTERPRISE_HEADERS]
    values = _enterprise()
    sheet.append([values[field] for _, field, _ in ENTERPRISE_HEADERS])
    assert normalized_rows(workbook) == [values]


def test_enterprise_import_stages_then_confirms_all_fields_atomically(tmp_db, tmp_path):
    """预览阶段不写正式表，确认后一次性写入全部企业字段并关联来源批次。"""
    workflow = _workflow(tmp_db, tmp_path)
    row = _enterprise()
    batch = workflow.parse_enterprises_and_stage("企业.xlsx", b"enterprise-source", [row])
    preview = workflow.preview_enterprises(batch["id"])
    assert preview["summary"] == {"new_enterprise": 1, "blocking": 0}

    connection = db_conn(tmp_db)
    assert connection.execute("SELECT COUNT(*) FROM enterprise").fetchone()[0] == 0
    connection.close()

    result = workflow.confirm_enterprises(batch["id"])
    assert result["enterprise_count"] == 1
    connection = db_conn(tmp_db)
    stored = connection.execute(
        "SELECT name,credit_code,enterprise_type,district,qualifications,contact_person,contact_phone,address,note "
        "FROM enterprise"
    ).fetchone()
    audit = connection.execute(
        "SELECT source_batch FROM audit_log WHERE object_type='enterprise' AND action='create'"
    ).fetchone()
    connection.close()
    assert dict(stored) == row
    assert audit["source_batch"] == str(batch["id"])


def test_duplicate_enterprise_blocks_entire_confirmation(tmp_db, tmp_path):
    """数据库已有或文件内重复的信用代码必须阻断批次，不能产生半批数据。"""
    workflow = _workflow(tmp_db, tmp_path)
    first = workflow.parse_enterprises_and_stage("首次.xlsx", b"first", [_enterprise()])
    workflow.confirm_enterprises(first["id"])
    duplicate = workflow.parse_enterprises_and_stage(
        "重复.xlsx", b"duplicate", [_enterprise(), _enterprise(code="91320583ENTERPRISE02", name="另一企业")]
    )
    preview = workflow.preview_enterprises(duplicate["id"])
    assert preview["summary"] == {"new_enterprise": 1, "blocking": 1}
    with pytest.raises(workflow.ConfirmationBlocked):
        workflow.confirm_enterprises(duplicate["id"])
    connection = db_conn(tmp_db)
    assert connection.execute("SELECT COUNT(*) FROM enterprise").fetchone()[0] == 1
    connection.close()


def test_unknown_enterprise_dictionary_value_is_visible_in_preview(tmp_db, tmp_path):
    """企业类型或区镇无法映射到系统字典时，预览必须直接显示字段错误。"""
    workflow = _workflow(tmp_db, tmp_path)
    row = _enterprise(code="91320583ENTERPRISE03", name="字典异常企业")
    row["district"] = "不存在的区镇"
    batch = workflow.parse_enterprises_and_stage("字典异常.xlsx", b"bad-dict", [row])
    preview = workflow.preview_enterprises(batch["id"])
    assert preview["summary"] == {"new_enterprise": 0, "blocking": 1}
    assert preview["rows"][0]["conclusion"] == "field_error"
    assert "区镇不在系统启用选项中" in preview["rows"][0]["error"]


def test_enterprise_modal_exposes_template_preview_and_confirm_flow():
    """前端新增企业弹窗必须接入企业模板、上传预览和确认接口。"""
    root = Path(__file__).resolve().parents[1]
    html = (root / "static" / "index.html").read_text(encoding="utf-8")
    # 导入流程已经拆到通用组件，契约覆盖入口与组件而不是单一入口文件。
    script = "\n".join(
        (root / "static" / path).read_text(encoding="utf-8")
        for path in ("app.js", "components/importer.js")
    )
    assert 'data-mtab="excel"' in html
    assert 'kind === "enterprise" ? "/enterprise-import" : "/import"' in script
    assert '"/enterprise-import"' in script
    assert '"/api/enterprise-template"' in script
    assert "btn-confirm-enterprise-import" in script


def test_enterprise_import_http_api_previews_and_confirms(client, tmp_db, tmp_path, monkeypatch):
    """真实 HTTP 路由能够接收工作簿、返回预览并执行人工确认。"""
    monkeypatch.setattr(app, "IMPORT_ARCHIVE_DIR", str(tmp_path / "http-imports"))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "企业台账"
    sheet.append([label + ("*" if required else "") for label, _, required in ENTERPRISE_HEADERS])
    row = _enterprise(code="91320583ENTERPRISE99", name="HTTP导入企业")
    sheet.append([row[field] for _, field, _ in ENTERPRISE_HEADERS])
    stream = io.BytesIO()
    workbook.save(stream)

    status, staged = client.request("POST", "/api/enterprise-import", {
        "filename": "HTTP企业导入.xlsx",
        "data": base64.b64encode(stream.getvalue()).decode("ascii"),
    })
    assert status == 200
    assert staged["preview"]["summary"] == {"new_enterprise": 1, "blocking": 0}

    status, confirmed = client.request("POST", f"/api/enterprise-import/{staged['id']}/confirm")
    assert status == 200 and confirmed["enterprise_count"] == 1
    connection = db_conn(tmp_db)
    assert connection.execute(
        "SELECT name FROM enterprise WHERE credit_code=?", (row["credit_code"],)
    ).fetchone()["name"] == row["name"]
    connection.close()
