#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导入模板下载契约：安装包资源完整，模板缺失时也能在当前进程内生成。"""

import io
from pathlib import Path

from openpyxl import load_workbook

import app
from import_excel import normalized_rows
from make_template import main as make_project_template
from release_tools.release_manifest import application_data_sources


class TemplateResponse:
    """记录模板接口写出的状态、响应头和二进制内容。"""

    def __init__(self):
        self.status = None
        self.headers = {}
        self.wfile = io.BytesIO()

    def send_response(self, status):
        self.status = status

    def send_header(self, name, value):
        self.headers[name] = value

    def end_headers(self):
        pass

    def _err(self, status, message):
        self.status = status
        self.error = message


def test_release_package_contains_import_template():
    """正式安装包必须直接携带模板，避免安装后首次下载再临时生成。"""
    project_root = Path(app.BASE_DIR)
    packaged = {source.name for source, _ in application_data_sources(project_root)}
    assert "导入模板.xlsx" in packaged


def test_template_api_generates_xlsx_without_starting_another_application(tmp_path, monkeypatch):
    """模板文件缺失时直接生成 Excel，并按附件下载响应返回给浏览器。"""
    monkeypatch.setattr(app, "BASE_DIR", str(tmp_path))
    response = TemplateResponse()

    app.Handler._api_template(response, "GET", [], {})

    assert response.status == 200
    assert response.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["Content-Disposition"] == 'attachment; filename="import_template.xlsx"'
    assert int(response.headers["Content-Length"]) == len(response.wfile.getvalue())
    workbook = load_workbook(io.BytesIO(response.wfile.getvalue()), read_only=True)
    assert workbook.sheetnames == ["项目台账", "填写说明"]


def test_generated_project_template_can_be_parsed_with_required_markers(tmp_path):
    """官方模板的必填星号必须与解析器兼容，且全部业务列都进入标准行。"""
    path = tmp_path / "import_template.xlsx"
    make_project_template(path)
    workbook = load_workbook(path)
    sheet = workbook["项目台账"]
    assert [cell.value for cell in sheet[1]][:10] == [
        "企业名称*", "统一社会信用代码*", "企业类型", "区镇", "资质", "企业联系人",
        "企业联系电话", "企业地址", "项目名称*", "项目编号/文号*",
    ]
    sheet.append([
        "测试企业", "91320000IMPORT0001", "其他", "开发区", "高新资质", "张三",
        "0512-12345678", "测试路1号", "测试项目", "IMPORT-001", "苏州市级", "创新联合体",
        100, "2026-01-01", "2026-12-31", "已立项", 1, "李四", "13800000000", "完整字段",
    ])
    workbook.save(path)

    rows = normalized_rows(load_workbook(path, data_only=True))
    assert rows == [{
        "enterprise_name": "测试企业",
        "credit_code": "91320000IMPORT0001",
        "enterprise_type": "其他",
        "district": "开发区",
        "qualifications": "高新资质",
        "enterprise_contact_person": "张三",
        "enterprise_contact_phone": "0512-12345678",
        "enterprise_address": "测试路1号",
        "project_name": "测试项目",
        "project_no": "IMPORT-001",
        "level": "苏州市级",
        "category": "创新联合体",
        "total_amount": 100.0,
        "start_date": "2026-01-01",
        "end_date": "2026-12-31",
        "stage": "已立项",
        "match_ratio": 1.0,
        "leader": "李四",
        "project_contact_phone": "13800000000",
        "project_note": "完整字段",
    }]
