#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成 Excel 导入模板（导入模板.xlsx）：单表「项目台账」= 企业字段 + 项目字段，导入时软件自动拆分归仓。

用法：python make_template.py   → 生成 ./导入模板.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = "导入模板.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1A5FB4")   # 项目字段（深蓝底）
ENT_FILL = PatternFill("solid", fgColor="2E7D32")     # 企业字段（深绿底，与白字对比清晰）
HEADER_FONT = Font(color="FFFFFF", bold=True)
REQ_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF7E6")
CENTER = Alignment(horizontal="center", vertical="center")

# (表头, 是否必填, 归属: 企业/项目)
HEADERS = [
    ("企业名称", True, "企业"),
    ("统一社会信用代码", False, "企业"),
    ("企业类型", False, "企业"),
    ("区镇", False, "企业"),
    ("资质", False, "企业"),
    ("企业联系人", False, "企业"),
    ("企业联系电话", False, "企业"),
    ("企业地址", False, "企业"),
    ("项目名称", True, "项目"),
    ("项目编号/文号", False, "项目"),
    ("层级", False, "项目"),
    ("类型", False, "项目"),
    ("总金额（万元）", False, "项目"),
    ("开始日期", False, "项目"),
    ("结束日期", False, "项目"),
    ("当前阶段", False, "项目"),
    ("配套比例", False, "项目"),
    ("项目负责人", False, "项目"),
    ("联系人手机号", False, "项目"),
    ("备注", False, "项目"),
]

GUIDE = """科技项目台账 - Excel 导入模板（自动拆分类）

【使用方法】
1. 下载本文件，只填「项目台账」这一个表。
2. 每一行 = 一个项目：左侧填企业信息（承担企业），右侧填项目信息。
3. 填好后：在系统点「新增项目 → Excel 批量导入」，把文件拖进虚线框（或点击选择），系统自动「颗粒归仓」：
   - 企业信息自动归入企业库；同一企业出现多行自动去重，不会重复建企业；
   - 每个项目自动关联到对应企业名下。
4. 也可命令行导入：python import_excel.py 你的文件.xlsx

【填写要求】
- 红色表头 = 必填列（企业名称、项目名称）。
- 企业名称 = 承担企业。同一企业在多行出现时，只有第一次需要填全企业信息；后续行企业名称一致即可（系统自动匹配已有企业，忽略其企业列）。
- 日期格式：YYYY-MM-DD（如 2026-01-01），也支持 2026/1/1、Excel 日期格式。
- 金额单位：万元（数字）。配套比例：如 1 表示 1:1（上级拨付:本级配套）。

【枚举取值】（必须用以下取值；如需新增，先在系统「配置」里加）
- 企业类型：高新技术企业 / 科技型中小企业 / 规上工业 / 其他
- 区镇：开发区 / 高新区 / 花桥 / 张浦 / 周市 / 陆家 / 巴城 / 千灯 / 周庄 / 淀山湖 / 锦溪
- 层级：国家级 / 省级 / 苏州市级 / 昆山本级
- 类型：科技成果转化 / 国际合作 / 创新联合体
- 阶段：申报中 / 已立项 / 实施中 / 待验收 / 已验收 / 绩效跟踪 / 已完结 / 中止 / 撤销

【说明】
- 资金拨付、项目节点不在本模板中，导入后在系统内逐条录入。
- 绿色表头（深绿） = 企业字段；蓝色表头（深蓝） = 项目字段。
"""


def main(output_path=OUT):
    """生成导入模板；调用方可指定输出路径，供安装版后端在当前进程内生成。"""
    wb = Workbook()

    # ---- 数据表优先：项目台账 放在最前（打开模板第一眼就是要填的表） ----
    ws = wb.active
    ws.title = "项目台账"
    ws.append([h[0] + ("*" if h[1] else "") for h in HEADERS])
    for col, (name, req, owner) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = ENT_FILL if owner == "企业" else HEADER_FILL
        cell.font = REQ_FONT if req else HEADER_FONT
        cell.alignment = CENTER
        width = 14 if owner == "企业" else 13
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 110

    # ---- 填写说明 放在最后 ----
    guide = wb.create_sheet("填写说明")
    guide.column_dimensions["A"].width = 66
    for i, line in enumerate(GUIDE.splitlines(), start=1):
        cell = guide.cell(row=i, column=1, value=line)
        if line.startswith("【") or line == "科技项目台账 - Excel 导入模板（自动拆分类）":
            cell.font = Font(bold=True, size=12)
        if line.strip().startswith("3."):
            cell.fill = NOTE_FILL
    guide.cell(row=1, column=1).font = Font(bold=True, size=14)

    wb.save(output_path)
    print(f"[OK] 已生成单表模板：{OUT}")
    print("sheet 顺序:", wb.sheetnames)
    print("列（深绿=企业字段，深蓝=项目字段）:")
    for name, req, owner in HEADERS:
        print(f"  [{owner}] {name}{'*' if req else ''}")


if __name__ == "__main__":
    main()
