#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 批量导入（单表自动拆分版）：读取一张总表（企业字段 + 项目字段），自动「颗粒归仓」：
   - 企业信息按名称自动去重后写入企业表（同一企业多行只建一次）；
   - 项目自动关联到对应企业名下。

用法：
  命令行：  python import_excel.py 你的文件.xlsx
  界面：    系统「新增项目 → Excel 批量导入」拖入/选择文件

表头（第一行，单表「项目台账」）：
  企业字段(绿)：企业名称*, 统一社会信用代码, 企业类型, 区镇, 资质, 企业联系人, 企业联系电话, 企业地址
  项目字段(蓝)：项目名称*, 项目编号/文号, 层级, 类型, 总金额（万元）, 开始日期, 结束日期,
                当前阶段, 配套比例, 项目负责人, 联系人手机号, 备注
返回：{"enterprise": {"ok": n, "errors": [...]}, "project": {"ok": n, "errors": [...]}}
"""

import datetime
import os
import re
import sys

# (表头 -> (归属, 字段))
COLS = {
    # 企业字段
    "企业名称": ("enterprise", "name"),
    "统一社会信用代码": ("enterprise", "credit_code"),
    "企业类型": ("enterprise", "enterprise_type"),
    "区镇": ("enterprise", "district"),
    "资质": ("enterprise", "qualifications"),
    "企业联系人": ("enterprise", "contact_person"),
    "企业联系电话": ("enterprise", "contact_phone"),
    "企业地址": ("enterprise", "address"),
    # 项目字段
    "项目名称": ("project", "name"),
    "项目编号/文号": ("project", "project_no"),
    "层级": ("project", "level"),
    "类型": ("project", "category"),
    "总金额（万元）": ("project", "total_amount"),
    "开始日期": ("project", "start_date"),
    "结束日期": ("project", "end_date"),
    "当前阶段": ("project", "stage"),
    "配套比例": ("project", "match_ratio"),
    "项目负责人": ("project", "leader"),
    "联系人手机号": ("project", "contact_phone"),
    "备注": ("project", "note"),
}

# 枚举校验（必须存在于配置 dict_item）
ENT_ENUMS = ["企业类型", "区镇"]
PROJ_ENUMS = ["层级", "类型"]
PROJECT_STAGES = {"申报中", "已立项", "实施中", "待验收", "已验收",
                  "绩效跟踪", "已完结", "中止", "撤销"}

DATA_SHEET_NAMES = ("项目台账", "项目")


def parse_date(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"^(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})日?$", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            datetime.date(y, mo, d)
        except ValueError:
            return None
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


def parse_number(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def clean_str(v):
    if isinstance(v, str):
        return v.strip() or None
    return v


def _pick_sheet(wb):
    for name in DATA_SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        if name != "填写说明":
            return wb[name]
    return None


def normalized_rows(wb):
    """把 Excel 解析为 G4 标准行；本函数只解析，不写任何数据库。"""
    sheet = _pick_sheet(wb)
    if sheet is None:
        return []
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    index = {name: position for position, name in enumerate(header) if name}

    def cell(row, name):
        position = index.get(name)
        return row[position] if position is not None and position < len(row) else None

    parsed = []
    for row in rows_iter:
        parsed.append({
            "enterprise_name": clean_str(cell(row, "企业名称")),
            "credit_code": clean_str(cell(row, "统一社会信用代码")),
            "enterprise_type": clean_str(cell(row, "企业类型")),
            "district": clean_str(cell(row, "区镇")),
            "project_name": clean_str(cell(row, "项目名称")),
            "project_no": clean_str(cell(row, "项目编号/文号")),
            "level": clean_str(cell(row, "层级")),
            "category": clean_str(cell(row, "类型")),
            "total_amount": parse_number(cell(row, "总金额（万元）")),
            "start_date": parse_date(cell(row, "开始日期")),
            "end_date": parse_date(cell(row, "结束日期")),
            "stage": clean_str(cell(row, "当前阶段")),
        })
    return parsed


def import_workbook(wb, conn, dict_map=None):
    """核心导入：单表自动拆分。dict_map: {dict_type: {value: True}}。返回结果 dict。"""
    if dict_map is None:
        rows = conn.execute("SELECT dict_type, value FROM dict_item WHERE is_active=1").fetchall()
        dict_map = {}
        for r in rows:
            dict_map.setdefault(r["dict_type"], set()).add(r["value"])

    result = {"enterprise": {"ok": 0, "errors": []}, "project": {"ok": 0, "errors": []}}

    sheet = _pick_sheet(wb)
    if sheet is None:
        return result

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return result
    idx = {h: i for i, h in enumerate(header) if h}

    def cell(row, header_name):
        i = idx.get(header_name)
        return row[i] if i is not None and i < len(row) else None

    # 已有企业缓存：name -> id（避免重复建企业）
    ent_cache = {r["name"]: r["id"] for r in conn.execute("SELECT id, name FROM enterprise")}
    existing_code = {r["credit_code"] for r in conn.execute(
        "SELECT credit_code FROM enterprise WHERE credit_code IS NOT NULL")}

    for row_no, row in enumerate(rows_iter, start=2):
        ent_name = clean_str(cell(row, "企业名称"))
        proj_name = clean_str(cell(row, "项目名称"))
        if not proj_name:
            result["project"]["errors"].append({"row": row_no, "reason": "项目名称必填"})
            continue
        if not ent_name:
            result["project"]["errors"].append({"row": row_no, "reason": "企业名称（承担企业）必填"})
            continue

        # ---- 1. 企业：命中缓存直接复用；否则校验并新建 ----
        ent_id = ent_cache.get(ent_name)
        if ent_id is None:
            ent_payload = {COLS[h][1]: _ent_val(h, cell(row, h)) for h in ENT_HEADERS}
            bad = []
            for h in ENT_ENUMS:
                v = ent_payload[COLS[h][1]]
                if v and v not in dict_map.get(_dict_type(h), set()):
                    bad.append(f"{h}『{v}』不在配置中")
            code = ent_payload.get("credit_code")
            if code and code in existing_code:
                bad.append("统一社会信用代码已存在")
            if bad:
                msg = "；".join(bad)
                result["enterprise"]["errors"].append({"row": row_no, "reason": msg})
                result["project"]["errors"].append({"row": row_no, "reason": f"企业『{ent_name}』未通过校验，项目未导入（{msg}）"})
                continue
            cols = list(ent_payload.keys())
            cur = conn.execute(
                f"INSERT INTO enterprise ({', '.join(cols)}) VALUES ({', '.join('?' * len(cols))})",
                [ent_payload[c] for c in cols])
            ent_id = cur.lastrowid
            ent_cache[ent_name] = ent_id
            if code:
                existing_code.add(code)
            result["enterprise"]["ok"] += 1

        # ---- 2. 项目：校验并插入 ----
        proj_payload = {}
        for h, (owner, field) in COLS.items():
            if owner != "project":
                continue
            raw = cell(row, h)
            if field in ("total_amount", "match_ratio"):
                val = parse_number(raw)
            elif field in ("start_date", "end_date"):
                val = parse_date(raw)
            else:
                val = clean_str(raw)
            proj_payload[field] = val
        bad = []
        for h in PROJ_ENUMS:
            v = proj_payload[COLS[h][1]]
            if v and v not in dict_map.get(_dict_type(h), set()):
                bad.append(f"{h}『{v}』不在配置中")
        if proj_payload.get("stage") and proj_payload["stage"] not in PROJECT_STAGES:
            bad.append(f"当前阶段『{proj_payload['stage']}』不是有效阶段")
        if proj_payload.get("start_date") is None and cell(row, "开始日期") not in (None, ""):
            bad.append("开始日期格式应为 YYYY-MM-DD")
        if proj_payload.get("end_date") is None and cell(row, "结束日期") not in (None, ""):
            bad.append("结束日期格式应为 YYYY-MM-DD")
        if bad:
            result["project"]["errors"].append({"row": row_no, "reason": "；".join(bad)})
            continue
        proj_payload["enterprise_id"] = ent_id
        cols = list(proj_payload.keys())
        conn.execute(
            f"INSERT INTO project ({', '.join(cols)}) VALUES ({', '.join('?' * len(cols))})",
            [proj_payload[c] for c in cols])
        result["project"]["ok"] += 1

    conn.commit()
    return result


ENT_HEADERS = ["企业名称", "统一社会信用代码", "企业类型", "区镇", "资质",
               "企业联系人", "企业联系电话", "企业地址"]


def _ent_val(h, raw):
    if h in ("企业类型", "区镇"):
        return clean_str(raw)
    return clean_str(raw)


def _dict_type(header_name):
    return {"企业类型": "enterprise_type", "区镇": "district",
            "层级": "level", "类型": "category"}[header_name]


def main():
    if len(sys.argv) < 2:
        print("用法: python import_excel.py 你的文件.xlsx")
        sys.exit(1)
    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"[错误] 文件不存在: {path}")
        sys.exit(1)
    from openpyxl import load_workbook
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    import app
    wb = load_workbook(path, data_only=True)
    conn = app.get_db()
    try:
        res = import_workbook(wb, conn)
    finally:
        conn.close()
    print(f"企业: 成功 {res['enterprise']['ok']} 条, 失败 {len(res['enterprise']['errors'])} 条")
    for e in res["enterprise"]["errors"]:
        print(f"  第{e['row']}行: {e['reason']}")
    print(f"项目: 成功 {res['project']['ok']} 条, 失败 {len(res['project']['errors'])} 条")
    for e in res["project"]["errors"]:
        print(f"  第{e['row']}行: {e['reason']}")
    sys.exit(0 if (res["enterprise"]["ok"] or res["project"]["ok"]) else 1)


if __name__ == "__main__":
    main()
