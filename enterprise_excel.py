#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""企业批量导入的 Excel 模板生成与标准化解析。"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ENTERPRISE_HEADERS = (
    ("企业名称", "name", True),
    ("统一社会信用代码", "credit_code", True),
    ("企业类型", "enterprise_type", False),
    ("区镇", "district", False),
    ("资质", "qualifications", False),
    ("联系人", "contact_person", False),
    ("联系电话", "contact_phone", False),
    ("地址", "address", False),
    ("备注", "note", False),
)


def build_template(output_path):
    """生成企业专用模板，字段顺序与新增企业表单保持一致。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "企业台账"
    sheet.append([label + ("*" if required else "") for label, _, required in ENTERPRISE_HEADERS])

    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    for column, (label, _, _) in enumerate(ENTERPRISE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        sheet.column_dimensions[get_column_letter(column)].width = max(14, len(label) * 2 + 4)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.zoomScale = 110

    guide = workbook.create_sheet("填写说明")
    guide.column_dimensions["A"].width = 72
    guide_rows = (
        "企业批量导入模板",
        "1. 每一行代表一家企业，企业名称和统一社会信用代码为必填项。",
        "2. 企业类型、区镇应使用系统“设置 - 配置”中已经启用的选项。",
        "3. 上传后系统先显示预览，只有全部数据通过检查后才能确认导入。",
        "4. 已存在或文件内重复的统一社会信用代码不会重复创建企业。",
    )
    for row, value in enumerate(guide_rows, start=1):
        guide.cell(row=row, column=1, value=value)
    guide.cell(row=1, column=1).font = Font(bold=True, size=14)
    workbook.save(output_path)


def normalized_rows(workbook):
    """把企业工作表转换为受控导入工作流使用的标准字段。"""
    sheet = workbook["企业台账"] if "企业台账" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(value).strip().removesuffix("*") if value is not None else "" for value in next(rows)]
    except StopIteration:
        return []
    positions = {name: index for index, name in enumerate(header) if name}

    def value_at(row, label):
        position = positions.get(label)
        if position is None or position >= len(row):
            return None
        value = row[position]
        return value.strip() or None if isinstance(value, str) else value

    normalized = []
    for row in rows:
        item = {field: value_at(row, label) for label, field, _ in ENTERPRISE_HEADERS}
        if any(value not in (None, "") for value in item.values()):
            normalized.append(item)
    return normalized
