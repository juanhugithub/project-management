#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""科技项目全生命周期台账系统 — 后端（Python 标准库，零第三方依赖）

双击 start.bat 启动，浏览器访问 http://127.0.0.1:8765
提供：静态文件服务 + JSON API（企业/项目/资金/节点/字典 的增删改查）
"""

import json
import os
import sqlite3
import threading
import webbrowser
import datetime
import sys
import subprocess
from io import BytesIO
from pathlib import Path
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, quote
from ledger.errors import DomainError
from ledger import queries, services
from ledger import field_mapping
from ledger import security
import usage_tracker
from runtime_paths import ensure_runtime_layout, get_runtime_paths
from version import get_version

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 安装目录只保存代码和资源；所有会变化的台账材料进入用户选择的运行目录。
RUNTIME_PATHS = get_runtime_paths()
DB_PATH = str(RUNTIME_PATHS.database)
IMPORT_ARCHIVE_DIR = str(RUNTIME_PATHS.import_archive)
SCHEMA_PATH = os.path.join(BASE_DIR, "schema.sql")
STATIC_DIR = os.path.join(BASE_DIR, "static")
APP_VERSION = get_version(BASE_DIR)
HOST = "127.0.0.1"
PORT = 8765
AUTH_ENABLED = os.environ.get("LEDGER_AUTH_ENABLED", "0") == "1"
INSTALL_CONFIG = os.environ.get("LEDGER_INSTALL_CONFIG", "")

# 各表允许写入的字段（白名单，防注入、防脏数据）
FIELDS = {
    "enterprise": ["name", "credit_code", "enterprise_type", "qualifications", "district",
                   "contact_person", "contact_phone", "address", "note"],
    "project": ["name", "project_no", "identity_status", "level", "category", "enterprise_id", "total_amount",
                "start_date", "end_date", "stage", "match_ratio", "leader", "contact_phone", "note"],
    "funding": ["project_id", "source_type", "amount", "batch", "plan_date", "actual_date", "status", "note"],
    "node": ["project_id", "node_type", "plan_date", "actual_date", "status", "has_major_change", "note"],
    "dict": ["dict_type", "value", "sort_order", "is_active"],
}

NUMERIC_FIELDS = {"total_amount", "amount", "match_ratio"}
INT_FIELDS = {"enterprise_id", "project_id", "has_major_change", "is_active", "sort_order"}

MIME = {
    ".html": "text/html; charset=utf-8",
    ".js": "application/javascript; charset=utf-8",
    ".css": "text/css; charset=utf-8",
    ".json": "application/json; charset=utf-8",
    # SVG 必须使用图像媒体类型返回，否则 Chromium 会把它视为普通二进制文件并拒绝渲染。
    ".svg": "image/svg+xml; charset=utf-8",
    ".ico": "image/x-icon",
}


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    ensure_runtime_layout(RUNTIME_PATHS)
    if not os.path.exists(DB_PATH):
        conn = get_db()
        with open(SCHEMA_PATH, "r", encoding="utf-8") as f:
            conn.executescript(f.read())
        conn.commit()
        conn.close()
    # 已存在的正式库绝不在应用启动时执行 SQL。结构升级仅能由显式迁移入口完成。
    # 企业列表的查询索引属于运行性能基础设施，不改变业务数据，启动时可幂等创建。
    conn = get_db()
    conn.executescript("""
        CREATE INDEX IF NOT EXISTS idx_enterprise_active_id
            ON enterprise(is_deleted, id DESC);
        CREATE INDEX IF NOT EXISTS idx_enterprise_name
            ON enterprise(name);
        CREATE INDEX IF NOT EXISTS idx_enterprise_district
            ON enterprise(district, is_deleted);
        CREATE INDEX IF NOT EXISTS idx_project_active_id
            ON project(is_deleted, id DESC);
        CREATE INDEX IF NOT EXISTS idx_project_stage
            ON project(stage, is_deleted);
    """)
    conn.commit()
    conn.close()


def clean_row(row):
    """sqlite3.Row -> dict"""
    return dict(row) if row is not None else None


def clean_payload(body, table):
    """只保留白名单字段；空字符串转 None；数值/整数字段做类型转换。"""
    out = {}
    allowed = FIELDS[table]
    for k, v in (body or {}).items():
        if k not in allowed:
            continue
        if isinstance(v, str):
            v = v.strip()
            if v == "":
                v = None
        elif k in NUMERIC_FIELDS and v is not None:
            try:
                v = float(v)
            except (TypeError, ValueError):
                v = None
        elif k in INT_FIELDS and v is not None:
            try:
                v = int(v)
            except (TypeError, ValueError):
                v = None
        out[k] = v
    return out


class Handler(BaseHTTPRequestHandler):
    server_version = "ProjectLedger/1.0"

    # ---------- 基础工具 ----------
    def _read_body(self):
        """每个请求只读取一次请求体，使鉴权拒绝后的连接也能正常收尾。"""
        if hasattr(self, "_parsed_request_body"):
            return self._parsed_request_body
        length = int(self.headers.get("Content-Length", 0))
        if not length:
            self._parsed_request_body = {}
            return self._parsed_request_body
        raw = self.rfile.read(length)
        try:
            self._parsed_request_body = json.loads(raw.decode("utf-8"))
        except Exception:
            self._parsed_request_body = {}
        return self._parsed_request_body

    def _send(self, status, obj, headers=None):
        data = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(data)

    def _send_file(self, data, filename, content_type):
        """发送内存中的导出文件，文件不写入安装目录或数据库目录。"""
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _ok(self, obj):
        self._send(200, obj)

    def _err(self, status, message):
        self._send(status, {"error": message})

    def _request_user(self):
        """从 HttpOnly 会话 Cookie 提取当前用户，API 不接受客户端伪造身份字段。"""
        raw_cookie = self.headers.get("Cookie", "")
        for item in raw_cookie.split(";"):
            key, sep, value = item.strip().partition("=")
            if key == "ledger_session" and sep:
                return security.user_for_token(value)
        return None

    def _require_role(self, *roles):
        """统一角色门禁：未登录 401，已登录但越权 403。"""
        if not self.current_user:
            self._err(401, "请先登录")
            return False
        if self.current_user["role"] not in roles:
            self._err(403, "当前角色无此操作权限")
            return False
        return True

    def _is_sensitive_action(self, resource, method, parts):
        """归档、恢复、导入是高影响操作，仅管理员可执行。"""
        return (
            resource == "config" and method == "PUT"
            or resource == "import" and method == "POST"
            or method == "POST" and len(parts) > 1 and parts[-1] == "restore"
        )

    # ---------- 路由 ----------
    def do_GET(self):
        self._route("GET")

    def do_POST(self):
        self._route("POST")

    def do_PUT(self):
        self._route("PUT")

    def do_DELETE(self):
        self._route("DELETE")

    def _route(self, method):
        parsed = urlparse(self.path)
        path = parsed.path.rstrip("/") or "/"
        parts = [p for p in path.split("/") if p]
        qs = parse_qs(parsed.query)

        # 静态文件 / 首页
        if method == "GET" and (path == "/" or path.startswith("/static/") or "." in parts[-1] if parts else False):
            self._serve_static(path)
            return

        # API
        if parts and parts[0] == "api":
            self._api(method, parts[1:], qs)
        else:
            self._serve_static("/")

    def _serve_static(self, path):
        if path == "/" or path == "":
            path = "/index.html"
        if path.startswith("/static/"):
            rel = path[len("/static/"):]
        else:
            rel = path.lstrip("/")
        # 防目录穿越
        file_path = os.path.normpath(os.path.join(STATIC_DIR, rel))
        if not file_path.startswith(STATIC_DIR):
            self._err(403, "forbidden")
            return
        if not os.path.isfile(file_path):
            self._err(404, "not found")
            return
        ext = os.path.splitext(file_path)[1].lower()
        content_type = MIME.get(ext, "application/octet-stream")
        with open(file_path, "rb") as f:
            data = f.read()
        if rel == "index.html":
            data = data.replace(b"__APP_VERSION__", APP_VERSION.encode("ascii"))
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-cache" if rel == "index.html" else "public, max-age=31536000, immutable")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    # ---------- API 分发 ----------
    def _api(self, method, parts, qs):
        # 必须先消费写请求体：若鉴权直接返回 401/403 而残留请求体，Windows 关闭
        # 套接字时会发送 RST，客户端便可能在已记录 403 后仍收到 ConnectionAbortedError。
        if method in ("POST", "PUT", "DELETE"):
            self._read_body()
        if not parts:
            self._err(404, "no api resource")
            return
        resource = parts[0]

        # 登录是唯一不需要既有会话的 API，其余读写一律先做身份认证。
        if resource == "auth":
            self._api_auth(method, parts[1:])
            return
        self.current_user = self._request_user() if AUTH_ENABLED else {"username": "local-user", "role": "admin", "district_scope": None}
        if AUTH_ENABLED and not self.current_user:
            self._err(401, "请先登录")
            return
        if resource == "usage" and method == "POST":
            body = self._read_body()
            usage_tracker.record(RUNTIME_PATHS, body.get("module"), body.get("action", "view"))
            self._ok({"recorded": True})
            return
        if resource == "usage" and method == "GET":
            self._ok(usage_tracker.summary(RUNTIME_PATHS))
            return
        if resource == "update":
            self._api_update(method, parts[1:])
            return
        if self._is_sensitive_action(resource, method, parts[1:]) and not self._require_role("admin"):
            return
        # 查阅员绝不写入；编辑员可做日常台账维护，但不能越过上面的敏感操作门禁。
        if method in ("POST", "PUT", "DELETE") and not self._require_role("admin", "editor"):
            return
        operator_token = security.set_current_operator(self.current_user["username"])
        try:
            self._dispatch_api(method, resource, parts[1:], qs)
        finally:
            security.reset_current_operator(operator_token)

    def _dispatch_api(self, method, resource, parts, qs):

        if resource == "dict":
            self._api_dict(method, parts, qs)
        elif resource == "enterprises":
            self._api_enterprise(method, parts, qs)
        elif resource == "projects":
            self._api_project(method, parts, qs)
        elif resource == "fundings":
            self._api_funding(method, parts, qs)
        elif resource == "nodes":
            self._api_node(method, parts, qs)
        elif resource == "reminders":
            self._api_reminders(method, parts, qs)
        elif resource in ("stats", "statistics"):
            self._api_stats(method, parts, qs)
        elif resource == "funding-check":
            self._api_funding_check(method, parts, qs)
        elif resource == "import":
            self._api_import(method, parts, qs)
        elif resource == "enterprise-import":
            self._api_enterprise_import(method, parts, qs)
        elif resource == "template":
            self._api_template(method, parts, qs)
        elif resource == "enterprise-template":
            self._api_enterprise_template(method, parts, qs)
        elif resource == "dashboard":
            self._api_dashboard(method, parts, qs)
        elif resource == "funding-plan":
            self._api_funding_plan(method, parts, qs)
        elif resource == "field-mapping":
            self._api_field_mapping(method, parts, qs)
        elif resource == "config":
            self._api_config(method, parts, qs)
        elif resource == "export":
            self._api_export(method, parts, qs)
        else:
            self._err(404, "unknown resource")

    def _api_auth(self, method, parts):
        """最小本地登录：成功后仅以 HttpOnly Cookie 建立会话。"""
        if method != "POST" or parts != ["login"]:
            self._err(405, "method not allowed")
            return
        body = self._read_body()
        token, user = security.login(body.get("username"), body.get("password"))
        if not token:
            self._err(401, "用户名或密码错误")
            return
        self._send(200, {"user": user}, {"Set-Cookie": f"ledger_session={token}; HttpOnly; SameSite=Strict; Path=/"})

    def _api_update(self, method, parts):
        """网页更新入口：复用安装版更新器，不下载或修改数据库。"""
        if method != "GET" and not (method == "POST" and parts == ["apply"]):
            self._err(405, "method not allowed"); return
        try:
            import installed_updater
            config_root = Path(INSTALL_CONFIG).parent if INSTALL_CONFIG else RUNTIME_PATHS.config
            manifest = os.environ.get("LEDGER_UPDATE_MANIFEST_URL") or installed_updater.configured_manifest_url(config_root)
            if method == "GET":
                result = installed_updater.check_installed_update(manifest, config_root)
                release = result["release"]
                self._ok({"current_version": result["current_version"], "release_version": release.version,
                          "update_available": result["update_available"], "notes": list(release.notes)})
                return
            # current-install.json 只记录当前版本和更新清单地址；程序、数据目录由
            # 安装器独立写入 install_locations.json。两者必须按各自契约读取，否则
            # 已安装版本点击更新时会因为 current-install.json 没有 program_root 而失败。
            current_install_path = Path(INSTALL_CONFIG) if INSTALL_CONFIG else config_root / "current-install.json"
            install_locations_path = config_root / "install_locations.json"
            install_locations = json.loads(install_locations_path.read_text(encoding="utf-8"))
            program_root = Path(install_locations["program_root"])
            data_root = Path(install_locations["data_root"])
            def apply_and_restart():
                installed_updater.apply_installed_update(manifest, program_root, data_root, config_root)
                current = json.loads(current_install_path.read_text(encoding="utf-8"))
                # 安装器完成后会把 current_version 切换到新版本；重启路径沿用安装位置
                # 配置中的 program_root，不能再次从版本配置中读取不存在的路径字段。
                new_exe = program_root / current["current_version"] / "项目台账" / "项目台账.exe"
                subprocess.Popen([str(new_exe), "--resident"], close_fds=True)
                self.server.shutdown()
            threading.Thread(target=apply_and_restart, daemon=True).start()
            self._ok({"started": True})
        except Exception as error:
            self._err(409, str(error))

    # ---------- 归档工具 ----------
    def _archived_years(self, conn):
        row = conn.execute("SELECT value FROM system_config WHERE key='archived_years'").fetchone()
        if not row or not row["value"]:
            return set()
        return {y.strip() for y in row["value"].split(",") if y.strip()}

    def _is_archived_project(self, conn, project_id):
        if not project_id:
            return False
        row = conn.execute("SELECT start_date FROM project WHERE id=?", (project_id,)).fetchone()
        if not row or not row["start_date"]:
            return False
        return row["start_date"][:4] in self._archived_years(conn)

    # ---------- 首页工作台 ----------
    def _api_dashboard(self, method, parts, qs):
        if method != "GET":
            self._err(405, "method not allowed"); return
        conn = get_db()
        try:
            totals = queries.dashboard(conn)
            project_count = conn.execute("SELECT COUNT(*) c FROM project").fetchone()["c"]
            enterprise_count = conn.execute("SELECT COUNT(*) c FROM enterprise").fetchone()["c"]
            funded_total = conn.execute("SELECT COALESCE(SUM(amount),0) a FROM funding WHERE status='已到账'").fetchone()["a"]
            plan_total = conn.execute("SELECT COALESCE(SUM(amount),0) a FROM funding WHERE plan_date IS NOT NULL").fetchone()["a"]
            overdue_funding = conn.execute(
                "SELECT COUNT(*) c, COALESCE(SUM(amount),0) a FROM funding "
                "WHERE plan_date IS NOT NULL AND plan_date < date('now','localtime') "
                "AND (actual_date IS NULL OR actual_date='')").fetchone()
            overdue_nodes = conn.execute(
                "SELECT COUNT(*) c FROM node WHERE status!='已完成' AND plan_date IS NOT NULL "
                "AND julianday(plan_date) < julianday(date('now','localtime'))").fetchone()["c"]
            # 工作台按季度观察节点，避免只看 30 天造成治理视野过短。
            due90_nodes = conn.execute(
                "SELECT COUNT(*) c FROM node WHERE status!='已完成' AND plan_date IS NOT NULL "
                "AND julianday(plan_date) BETWEEN julianday(date('now','localtime')) "
                "AND julianday(date('now','localtime'))+90").fetchone()["c"]
            by_level = [clean_row(r) for r in conn.execute(
                "SELECT COALESCE(level,'未设置') AS key, COUNT(*) AS count FROM project "
                "GROUP BY level ORDER BY count DESC").fetchall()]
            by_category = [clean_row(r) for r in conn.execute(
                "SELECT COALESCE(category,'未设置') AS key, COUNT(*) AS count FROM project "
                "GROUP BY category ORDER BY count DESC").fetchall()]
            self._ok({
                **totals,
                "project_count": project_count,
                "enterprise_count": enterprise_count,
                "funded_total": round(funded_total, 2),
                "plan_total": round(plan_total, 2),
                "overdue_funding_count": overdue_funding["c"],
                "overdue_funding_amount": round(overdue_funding["a"], 2),
                "overdue_nodes": overdue_nodes,
                "due90_nodes": due90_nodes,
                "by_level": by_level,
                "by_category": by_category,
            })
        finally:
            conn.close()

    # ---------- 资金拨付执行度 ----------
    def _api_funding_plan(self, method, parts, qs):
        if method != "GET":
            self._err(405, "method not allowed"); return
        conn = get_db()
        try:
            rows = conn.execute(
                "SELECT f.id, f.project_id, f.source_type, f.amount, f.plan_date, f.actual_date, f.status, "
                "p.name AS project_name, p.level AS project_level, p.start_date "
                "FROM funding f JOIN project p ON f.project_id=p.id "
                "WHERE f.plan_date IS NOT NULL ORDER BY f.plan_date, f.id").fetchall()
            today = datetime.date.today().isoformat()
            items = []
            summary = {"total_plan": 0.0, "total_paid": 0.0, "overdue_count": 0, "overdue_amount": 0.0}
            for r in rows:
                d = clean_row(r)
                d["is_overdue"] = bool(d["plan_date"] and d["plan_date"] < today and not d["actual_date"])
                items.append(d)
                summary["total_plan"] += d["amount"] or 0
                if d["status"] == "已到账":
                    summary["total_paid"] += d["amount"] or 0
                if d["is_overdue"]:
                    summary["overdue_count"] += 1
                    summary["overdue_amount"] += d["amount"] or 0
            summary["total_plan"] = round(summary["total_plan"], 2)
            summary["total_paid"] = round(summary["total_paid"], 2)
            summary["overdue_amount"] = round(summary["overdue_amount"], 2)
            summary["execution_rate"] = round(summary["total_paid"] / summary["total_plan"], 4) if summary["total_plan"] else 0
            self._ok({"items": items, "summary": summary})
        finally:
            conn.close()

    # ---------- 系统配置 ----------
    def _api_config(self, method, parts, qs):
        conn = get_db()
        try:
            if method == "GET":
                rows = conn.execute("SELECT key, value FROM system_config").fetchall()
                out = {}
                for r in rows:
                    if r["key"] == "archived_years":
                        out[r["key"]] = [y for y in (r["value"] or "").split(",") if y.strip()]
                    elif r["key"] == "ui_texts":
                        try:
                            out[r["key"]] = json.loads(r["value"] or "{}")
                        except json.JSONDecodeError:
                            out[r["key"]] = {}
                    else:
                        out[r["key"]] = r["value"]
                self._ok(out)
            elif method == "PUT":
                body = self._read_body()
                years = (body or {}).get("archived_years")
                ui_texts = (body or {}).get("ui_texts")
                if years is not None and not isinstance(years, list):
                    self._err(400, "archived_years must be list"); return
                if years is not None:
                    services.set_archived_years(conn, years, (body or {}).get("reason"))
                if ui_texts is not None:
                    if not isinstance(ui_texts, dict) or any(not isinstance(k, str) or not isinstance(v, str) for k, v in ui_texts.items()):
                        self._err(400, "ui_texts must be an object of strings"); return
                    conn.execute("INSERT INTO system_config(key,value) VALUES('ui_texts',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                                 (json.dumps(ui_texts, ensure_ascii=False),))
                    conn.commit()
                self._ok({"saved": True})
            else:
                self._err(405, "method not allowed")
        finally:
            conn.close()

    # ---------- 下载导入模板 ----------
    def _api_template(self, method, parts, qs):
        if method != "GET":
            self._err(405, "method not allowed"); return
        path = os.path.join(BASE_DIR, "导入模板.xlsx")
        if not os.path.exists(path):
            # 安装版中的 sys.executable 是“项目台账.exe”，不能用它执行 Python 脚本，
            # 否则会再次启动项目网页。模板缺失时直接调用生成函数，并明确写入接口读取路径。
            import make_template
            make_template.main(path)
        if not os.path.exists(path):
            self._err(500, "模板生成失败"); return
        with open(path, "rb") as f:
            data = f.read()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", 'attachment; filename="import_template.xlsx"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _api_enterprise_template(self, method, parts, qs):
        """下载企业专用模板；模板字段与新增企业表单保持一致。"""
        if method != "GET":
            self._err(405, "method not allowed"); return
        from enterprise_excel import build_template
        path = os.path.join(BASE_DIR, "企业导入模板.xlsx")
        if not os.path.exists(path):
            build_template(path)
        with open(path, "rb") as file:
            data = file.read()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", 'attachment; filename="enterprise_import_template.xlsx"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    # ---------- Excel 导入 ----------
    def _api_import(self, method, parts, qs):
        if method == "GET" and len(parts) == 1:
            from imports.controlled import ImportWorkflow
            try:
                workflow = ImportWorkflow(DB_PATH, IMPORT_ARCHIVE_DIR, apply_schema=False)
                self._ok(workflow.preview(int(parts[0])))
            except (DomainError, ValueError, sqlite3.Error) as exc:
                self._err(400, str(exc))
            return
        if method == "POST" and len(parts) == 2 and parts[1] == "confirm":
            from imports.controlled import ImportWorkflow
            try:
                workflow = ImportWorkflow(DB_PATH, IMPORT_ARCHIVE_DIR, apply_schema=False)
                self._ok(workflow.confirm(int(parts[0])))
            except DomainError as exc:
                self._err(409, str(exc))
            except (ValueError, sqlite3.Error) as exc:
                self._err(400, str(exc))
            return
        if method != "POST" or parts:
            self._err(405, "method not allowed"); return
        try:
            import base64
            import io
            from openpyxl import load_workbook
        except ImportError:
            self._err(500, "未安装 openpyxl，请先运行: pip install openpyxl"); return
        body = self._read_body()
        b64 = body.get("data") or ""
        try:
            raw = base64.b64decode(b64)
        except Exception:
            self._err(400, "文件内容无法解码"); return
        try:
            wb = load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as e:
            self._err(400, f"Excel 解析失败: {e}"); return
        try:
            from import_excel import normalized_rows
            from imports.controlled import ImportWorkflow
        except ImportError:
            self._err(500, "缺少 import_excel.py"); return
        try:
            workflow = ImportWorkflow(DB_PATH, IMPORT_ARCHIVE_DIR, apply_schema=False)
            result = workflow.parse_and_stage(body.get("name") or "upload.xlsx", raw, normalized_rows(wb), "excel-v1")
            result["preview"] = workflow.preview(result["id"])
        except (DomainError, sqlite3.Error) as e:
            self._err(400, f"导入暂存失败: {e}"); return
        self._ok(result)

    def _api_enterprise_import(self, method, parts, qs):
        """企业 Excel 受控导入：先暂存预览，再由页面明确确认入库。"""
        from imports.controlled import ImportWorkflow
        workflow = ImportWorkflow(DB_PATH, IMPORT_ARCHIVE_DIR, apply_schema=False)
        if method == "POST" and len(parts) == 2 and parts[1] == "confirm":
            try:
                self._ok(workflow.confirm_enterprises(int(parts[0])))
            except workflow.ConfirmationBlocked as error:
                self._err(409, str(error))
            except (DomainError, ValueError, sqlite3.Error) as error:
                self._err(400, str(error))
            return
        if method != "POST" or parts:
            self._err(405, "method not allowed"); return

        import base64
        import io
        from openpyxl import load_workbook
        from enterprise_excel import normalized_rows

        body = self._read_body()
        try:
            raw = base64.b64decode(body.get("data") or "")
            workbook = load_workbook(io.BytesIO(raw), data_only=True)
            rows = normalized_rows(workbook)
            if not rows:
                raise DomainError("企业模板中没有可导入的数据行")
            result = workflow.parse_enterprises_and_stage(body.get("filename") or "企业导入.xlsx", raw, rows)
            result["preview"] = workflow.preview_enterprises(result["id"])
            self._ok(result)
        except (DomainError, ValueError, sqlite3.Error) as error:
            self._err(400, str(error))

    # ---------- 字典 ----------
    def _api_dict(self, method, parts, qs):
        conn = get_db()
        try:
            if method == "GET":
                if parts and parts[0] == "types":
                    rows = conn.execute("SELECT DISTINCT dict_type FROM dict_item ORDER BY dict_type").fetchall()
                    self._ok([r["dict_type"] for r in rows])
                    return
                dict_type = qs.get("type", [None])[0]
                all_items = qs.get("all", ["0"])[0] == "1"
                if dict_type:
                    sql = "SELECT * FROM dict_item WHERE dict_type=?"
                    if not all_items:
                        sql += " AND is_active=1"
                    rows = conn.execute(sql + " ORDER BY sort_order, id", (dict_type,)).fetchall()
                    self._ok([clean_row(r) for r in rows])
                else:
                    rows = conn.execute(
                        "SELECT dict_type, value FROM dict_item WHERE is_active=1 ORDER BY dict_type, sort_order, id"
                    ).fetchall()
                    grouped = {}
                    for r in rows:
                        grouped.setdefault(r["dict_type"], []).append(r["value"])
                    self._ok(grouped)
            elif method == "POST" and not parts:
                body = clean_payload(self._read_body(), "dict")
                if not body.get("dict_type") or not body.get("value"):
                    self._err(400, "dict_type and value are required"); return
                dup = conn.execute(
                    "SELECT id FROM dict_item WHERE dict_type=? AND value=?",
                    (body["dict_type"], body["value"])).fetchone()
                if dup:
                    self._err(400, "该取值已存在"); return
                cur = conn.execute(
                    "INSERT INTO dict_item (dict_type, value, sort_order, is_active) VALUES (?,?,?,1)",
                    (body["dict_type"], body["value"], body.get("sort_order") or 0))
                conn.commit()
                row = conn.execute("SELECT * FROM dict_item WHERE id=?", (cur.lastrowid,)).fetchone()
                self._ok(clean_row(row))
            elif method == "PUT" and len(parts) == 1:
                did = int(parts[0])
                body = clean_payload(self._read_body(), "dict")
                if not body:
                    self._err(400, "no fields"); return
                sets = ", ".join(f"{k}=?" for k in body)
                conn.execute(f"UPDATE dict_item SET {sets} WHERE id=?", [*body.values(), did])
                conn.commit()
                row = conn.execute("SELECT * FROM dict_item WHERE id=?", (did,)).fetchone()
                self._ok(clean_row(row))
            elif method == "DELETE" and len(parts) == 1:
                did = int(parts[0])
                # 停用而非删除，保护历史引用
                conn.execute("UPDATE dict_item SET is_active=0 WHERE id=?", (did,))
                conn.commit()
                self._ok({"disabled": did})
            else:
                self._err(405, "method not allowed")
        except DomainError as exc:
            self._err(403 if '已归档' in str(exc) else 409 if '父项目' in str(exc) else 400, str(exc))
        except ValueError:
            self._err(400, "invalid id")
        finally:
            conn.close()

    # ---------- Excel 导出 ----------
    def _api_export(self, method, parts, qs):
        """按页面当前筛选条件生成 Excel；只读导出不改变业务数据。"""
        if method != "GET" or parts:
            self._err(405, "method not allowed"); return
        resource = qs.get("resource", [""])[0]
        try:
            from openpyxl import Workbook
        except ImportError:
            self._err(500, "未安装 openpyxl"); return
        conn = get_db()
        try:
            workbook = Workbook()
            sheet = workbook.active
            if resource == "enterprises":
                q = qs.get("q", [""])[0].strip()
                where = ["e.is_deleted=0"]; params = []
                if q:
                    where.append("(e.name LIKE ? OR e.credit_code LIKE ? OR e.district LIKE ? OR e.enterprise_type LIKE ?)")
                    like = f"%{q}%"; params.extend([like, like, like, like])
                rows = conn.execute(
                    "SELECT e.name, e.credit_code, e.enterprise_type, e.district, e.qualifications, "
                    "(SELECT COUNT(*) FROM project p WHERE p.enterprise_id=e.id AND p.is_deleted=0), "
                    "(SELECT COALESCE(SUM(p.total_amount),0) FROM project p WHERE p.enterprise_id=e.id AND p.is_deleted=0) "
                    f"FROM enterprise e WHERE {' AND '.join(where)} ORDER BY e.id DESC", params).fetchall()
                sheet.title = "企业结果"
                sheet.append(["企业名称", "统一社会信用代码", "企业类型", "区镇", "资质", "项目数", "累计金额(万元)"])
                for row in rows: sheet.append(list(row))
                filename = "企业当前结果.xlsx"
            elif resource == "reminders":
                days = int(qs.get("days", ["90"])[0])
                rows = conn.execute(
                    "SELECT p.name, p.level, n.node_type, n.plan_date, n.status, "
                    "(julianday(n.plan_date)-julianday(date('now','localtime'))) "
                    "FROM node n JOIN project p ON n.project_id=p.id "
                    "WHERE n.status!='已完成' AND n.plan_date IS NOT NULL "
                    "AND (julianday(n.plan_date)-julianday(date('now','localtime'))) <= ? "
                    "ORDER BY n.plan_date", (days,)).fetchall()
                sheet.title = "提醒结果"
                sheet.append(["项目名称", "层级", "节点类型", "计划时间", "状态", "剩余天数"])
                for row in rows: sheet.append(list(row))
                filename = "提醒当前结果.xlsx"
            else:
                self._err(400, "unknown export resource"); return
            for cell in sheet[1]: cell.font = cell.font.copy(bold=True)
            output = BytesIO(); workbook.save(output)
            self._send_file(output.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except (ValueError, TypeError):
            self._err(400, "invalid export parameters")
        finally:
            conn.close()

    def _api_field_mapping(self, method, parts, qs):
        """提供字段字典、候选映射和已确认行翻译，事实写入仍由人工确认流程完成。"""
        if not parts or (parts == ["dictionary"] and method != "GET") or (parts != ["dictionary"] and method != "POST"):
            self._err(405, "method not allowed"); return
        body = self._read_body()
        try:
            if parts == ["dictionary"]:
                self._ok(field_mapping.standard_fields())
            elif parts == ["suggest"]:
                self._ok(field_mapping.suggest_mapping(body.get("headers")))
            elif parts == ["translate"]:
                self._ok(field_mapping.translate_rows(body.get("rows"), body.get("mapping")))
            else:
                self._err(404, "unknown field mapping action")
        except field_mapping.FieldMappingError as error:
            self._err(400, str(error))

    # ---------- 提醒 ----------
    def _api_reminders(self, method, parts, qs):
        if method != "GET":
            self._err(405, "method not allowed"); return
        conn = get_db()
        try:
            days = int(qs.get("days", ["30"])[0])
            rows = conn.execute(
                "SELECT n.id, n.project_id, n.node_type, n.plan_date, n.actual_date, n.status, n.note, "
                "p.name AS project_name, p.level AS project_level, "
                "(julianday(n.plan_date) - julianday(date('now','localtime'))) AS days_left "
                "FROM node n JOIN project p ON n.project_id = p.id "
                "WHERE n.status != '已完成' AND n.plan_date IS NOT NULL "
                "AND (julianday(n.plan_date) - julianday(date('now','localtime'))) <= ? "
                "ORDER BY n.plan_date",
                (days,)
            ).fetchall()
            out = []
            for r in rows:
                d = clean_row(r)
                dl = d.get("days_left")
                if dl is None:
                    d["level"] = "later"
                elif dl < 0:
                    d["level"] = "overdue"
                elif dl <= 7:
                    d["level"] = "red"
                elif dl <= days:
                    d["level"] = "yellow"
                else:
                    d["level"] = "later"
                out.append(d)
            self._ok(out)
        except ValueError:
            self._err(400, "invalid days")
        finally:
            conn.close()

    # ---------- 统计 ----------
    def _api_stats(self, method, parts, qs):
        if method != "GET":
            self._err(405, "method not allowed"); return
        by = qs.get("by", ["category"])[0]
        conn = get_db()
        try:
            scope = self.current_user.get("district_scope")
            scope_sql = " AND e.district=?" if scope else ""
            scope_params = (scope,) if scope else ()
            if by == "district":
                rows = conn.execute(
                    "SELECT COALESCE(e.district,'未设置') AS key, COUNT(p.id) AS count, "
                    "COALESCE(SUM(p.total_amount),0) AS amount "
                    "FROM project p JOIN enterprise e ON p.enterprise_id=e.id "
                    "WHERE p.is_deleted=0 AND e.is_deleted=0" + scope_sql + " GROUP BY e.district ORDER BY count DESC",
                    scope_params,
                ).fetchall()
            elif by == "source":
                rows = conn.execute(
                    "SELECT f.source_type AS key, COUNT(*) AS count, COALESCE(SUM(f.amount),0) AS amount "
                    "FROM funding f JOIN project p ON f.project_id=p.id JOIN enterprise e ON p.enterprise_id=e.id "
                    "WHERE f.is_deleted=0 AND p.is_deleted=0 AND e.is_deleted=0" + scope_sql + " GROUP BY f.source_type ORDER BY amount DESC",
                    scope_params).fetchall()
            elif by == "enterprise":
                rows = conn.execute(
                    "SELECT COALESCE(e.name,'未关联') AS key, COUNT(p.id) AS count, "
                    "COALESCE(SUM(p.total_amount),0) AS amount "
                    "FROM project p LEFT JOIN enterprise e ON p.enterprise_id=e.id "
                    "GROUP BY p.enterprise_id ORDER BY amount DESC").fetchall()
            elif by == "year":
                rows = conn.execute(
                    "SELECT substr(p.start_date,1,4) AS key, COUNT(*) AS count, "
                    "COALESCE(SUM(p.total_amount),0) AS amount "
                    "FROM project p GROUP BY substr(p.start_date,1,4) ORDER BY key").fetchall()
            elif by == "stage":
                rows = conn.execute(
                    "SELECT p.stage AS key, COUNT(*) AS count, COALESCE(SUM(p.total_amount),0) AS amount "
                    "FROM project p GROUP BY p.stage ORDER BY count DESC").fetchall()
            else:
                col = by if by in ("level", "category") else "category"
                rows = conn.execute(
                    f"SELECT p.{col} AS key, COUNT(*) AS count, COALESCE(SUM(p.total_amount),0) AS amount "
                    f"FROM project p GROUP BY p.{col} ORDER BY count DESC").fetchall()
            out = []
            for r in rows:
                d = clean_row(r)
                d["key"] = d.get("key") or "未设置"
                out.append(d)
            self._ok(out)
        finally:
            conn.close()

    # ---------- 资金勾稽核对 ----------
    def _api_funding_check(self, method, parts, qs):
        if method != "GET":
            self._err(405, "method not allowed"); return
        conn = get_db()
        try:
            rows = conn.execute(
                "SELECT p.id, p.name, p.total_amount, p.match_ratio, "
                "COALESCE(SUM(CASE WHEN f.source_type='上级拨付' THEN f.amount ELSE 0 END),0) AS sum_up, "
                "COALESCE(SUM(CASE WHEN f.source_type='本级配套' THEN f.amount ELSE 0 END),0) AS sum_match, "
                "COALESCE(SUM(CASE WHEN f.source_type='本级自付' THEN f.amount ELSE 0 END),0) AS sum_self "
                "FROM project p LEFT JOIN funding f ON f.project_id=p.id "
                "GROUP BY p.id ORDER BY p.id").fetchall()
            out = []
            for r in rows:
                d = clean_row(r)
                d["sum_all"] = d["sum_up"] + d["sum_match"] + d["sum_self"]
                issues = []
                if d["total_amount"] is not None and abs(d["sum_all"] - d["total_amount"]) > 0.005:
                    issues.append("资金合计与项目总金额不一致")
                if d["match_ratio"] and d["sum_up"]:
                    expected = d["sum_up"] * d["match_ratio"]
                    d["match_expected"] = round(expected, 2)
                    if abs(d["sum_match"] - expected) > 0.005:
                        issues.append("本级配套与应配额不一致")
                d["issues"] = issues
                d["ok"] = len(issues) == 0
                out.append(d)
            self._ok(out)
        finally:
            conn.close()

    # ---------- 企业 ----------
    def _api_enterprise(self, method, parts, qs):
        conn = get_db()
        try:
            if method == "GET" and not parts:
                # 选项查询只返回项目表单需要的字段，避免把企业画像和统计字段带到前端。
                if qs.get("lookup", [""])[0] == "1":
                    rows = conn.execute(
                        "SELECT id, name, credit_code FROM enterprise "
                        "WHERE is_deleted=0 AND is_active=1 ORDER BY name, id"
                    ).fetchall()
                    self._ok([clean_row(r) for r in rows])
                    return
                # 传入 page/page_size 时启用分页协议；不传参数仍保留旧数组响应，兼容外部调用者。
                if not any(k in qs for k in ("page", "page_size", "q", "sort")):
                    rows = conn.execute(
                        "SELECT e.*, "
                        "(SELECT COUNT(*) FROM project p WHERE p.enterprise_id=e.id AND p.is_deleted=0) AS project_count, "
                        "(SELECT COALESCE(SUM(p.total_amount),0) FROM project p WHERE p.enterprise_id=e.id AND p.is_deleted=0) AS total_amount_sum "
                        "FROM enterprise e WHERE e.is_deleted=0 ORDER BY e.id DESC"
                    ).fetchall()
                    self._ok([clean_row(r) for r in rows])
                    return
                page = max(1, int(qs.get("page", ["1"])[0]))
                page_size = min(100, max(20, int(qs.get("page_size", ["50"])[0])))
                q = qs.get("q", [""])[0].strip()
                where = ["e.is_deleted=0"]
                params = []
                if q:
                    where.append("(e.name LIKE ? OR e.credit_code LIKE ? OR e.district LIKE ? OR e.enterprise_type LIKE ?)")
                    like = f"%{q}%"
                    params.extend([like, like, like, like])
                sort_map = {
                    "name": "e.name", "credit_code": "e.credit_code",
                    "enterprise_type": "e.enterprise_type", "district": "e.district",
                    "project_count": "project_count", "total_amount_sum": "total_amount_sum",
                }
                sort = sort_map.get(qs.get("sort", ["id"])[0], "e.id")
                direction = "ASC" if qs.get("direction", ["desc"])[0].lower() == "asc" else "DESC"
                where_sql = " AND ".join(where)
                total = conn.execute(f"SELECT COUNT(*) FROM enterprise e WHERE {where_sql}", params).fetchone()[0]
                offset = (page - 1) * page_size
                rows = conn.execute(
                    "SELECT e.*, "
                    "(SELECT COUNT(*) FROM project p WHERE p.enterprise_id=e.id AND p.is_deleted=0) AS project_count, "
                    "(SELECT COALESCE(SUM(p.total_amount),0) FROM project p WHERE p.enterprise_id=e.id AND p.is_deleted=0) AS total_amount_sum "
                    f"FROM enterprise e WHERE {where_sql} ORDER BY {sort} {direction}, e.id DESC LIMIT ? OFFSET ?",
                    params + [page_size, offset],
                ).fetchall()
                self._ok({"items": [clean_row(r) for r in rows], "total": total,
                          "page": page, "page_size": page_size,
                          "total_pages": (total + page_size - 1) // page_size})
            elif method == "GET" and len(parts) == 1:
                eid = int(parts[0])
                ent = conn.execute("SELECT * FROM enterprise WHERE id=? AND is_deleted=0", (eid,)).fetchone()
                if not ent:
                    self._err(404, "enterprise not found"); return
                projects = conn.execute(
                    "SELECT * FROM project WHERE enterprise_id=? AND is_deleted=0 ORDER BY id DESC", (eid,)
                ).fetchall()
                result = clean_row(ent)
                result["projects"] = [clean_row(r) for r in projects]
                self._ok(result)
            elif method == "POST" and not parts:
                payload = clean_payload(self._read_body(), "enterprise")
                if not payload.get("name"):
                    self._err(400, "name is required"); return
                self._ok(services.create(conn, "enterprise", payload))
            elif method == "PUT" and len(parts) == 1:
                eid = int(parts[0])
                payload = clean_payload(self._read_body(), "enterprise")
                if not payload:
                    self._err(400, "no fields"); return
                self._ok(services.update(conn, "enterprise", eid, payload))
            elif method == "DELETE" and len(parts) == 1:
                eid = int(parts[0])
                services.soft_delete(conn, "enterprise", eid, self._read_body().get("reason"))
                self._ok({"deleted": eid})
            elif method == "POST" and len(parts) == 2 and parts[1] == "restore":
                self._ok(services.restore(conn, "enterprise", int(parts[0]), self._read_body().get("reason")))
            elif method == "POST" and len(parts) == 2 and parts[1] in ("disable", "enable"):
                body = self._read_body()
                self._ok(services.set_enterprise_active(conn, int(parts[0]), parts[1] == "enable", body.get("reason")))
            else:
                self._err(405, "method not allowed")
        except DomainError as exc:
            self._err(403 if '已归档' in str(exc) else 409 if '父项目' in str(exc) else 400, str(exc))
        except ValueError:
            self._err(400, "invalid id")
        finally:
            conn.close()

    # ---------- 项目 ----------
    def _build_filter_cond(self, field, op, val):
        """高级筛选：字段/运算符/值 -> (sql片段, 参数)。非法组合返回 None。"""
        if val is None or (isinstance(val, str) and val.strip() == ""):
            return None
        colmap = {
            "name": "p.name", "project_no": "p.project_no", "level": "p.level",
            "category": "p.category", "stage": "p.stage", "enterprise_id": "p.enterprise_id",
            "enterprise_name": "e.name", "district": "e.district",
            "total_amount": "p.total_amount", "match_ratio": "p.match_ratio",
            "start_date": "p.start_date", "end_date": "p.end_date",
            "leader": "p.leader", "contact_phone": "p.contact_phone",
        }
        col = colmap.get(field)
        if not col:
            return None
        if op == "eq":
            return (f"{col}=?", val)
        if op == "contains":
            return (f"{col} LIKE ?", f"%{val}%")
        if op == "gt":
            return (f"{col} > ?", val)
        if op == "gte":
            return (f"{col} >= ?", val)
        if op == "lt":
            return (f"{col} < ?", val)
        if op == "lte":
            return (f"{col} <= ?", val)
        return None

    def _api_project(self, method, parts, qs):
        conn = get_db()
        try:
            if method == "GET" and not parts:
                where, params = [], []
                for key, col in [("level", "p.level"), ("category", "p.category"),
                                 ("stage", "p.stage"), ("enterprise_id", "p.enterprise_id")]:
                    v = qs.get(key, [None])[0]
                    if v:
                        where.append(f"{col}=?")
                        params.append(v)
                # 区镇（挂在承担企业上）
                v = qs.get("district", [None])[0]
                if v:
                    where.append("e.district=?")
                    params.append(v)
                q = qs.get("q", [None])[0]
                if q:
                    where.append("(p.name LIKE ? OR p.project_no LIKE ? OR e.name LIKE ?)")
                    like = f"%{q}%"
                    params += [like, like, like]
                # 高级筛选（JSON 数组：[{field, op, value}, ...] AND 组合）
                filters_raw = qs.get("filters", [None])[0]
                if filters_raw:
                    try:
                        for cond in json.loads(filters_raw):
                            built = self._build_filter_cond(
                                cond.get("field"), cond.get("op"), cond.get("value"))
                            if built:
                                where.append(built[0])
                                params.append(built[1])
                    except (ValueError, TypeError):
                        pass
                requested_district = qs.get("district", [None])[0]
                scoped_district = self.current_user.get("district_scope")
                if scoped_district and requested_district and requested_district != scoped_district:
                    self._ok([]); return
                project_filters = {
                    "level": qs.get("level", [None])[0], "category": qs.get("category", [None])[0],
                    "stage": qs.get("stage", [None])[0], "enterprise_id": qs.get("enterprise_id", [None])[0],
                    "district": scoped_district or requested_district, "query": qs.get("q", [None])[0]}
                if filters_raw:
                    try: project_filters["adv_filters"] = json.loads(filters_raw)
                    except (ValueError, TypeError): project_filters["adv_filters"] = []
                if "page" in qs or "page_size" in qs or "sort" in qs:
                    project_filters.update({"page": qs.get("page", ["1"])[0],
                                            "page_size": qs.get("page_size", ["50"])[0],
                                            "sort": qs.get("sort", ["id"])[0],
                                            "direction": qs.get("direction", ["desc"])[0]})
                self._ok(queries.project_list(conn, project_filters))
            elif method == "GET" and len(parts) == 1:
                pid = int(parts[0])
                result = queries.project_detail(conn, pid)
                if not result:
                    self._err(404, "project not found"); return
                self._ok(result)
            elif method == "POST" and not parts:
                payload = clean_payload(self._read_body(), "project")
                if not payload.get("name"):
                    self._err(400, "name is required"); return
                self._ok(services.create(conn, "project", payload))
            elif method == "PUT" and len(parts) == 1:
                pid = int(parts[0])
                if self._is_archived_project(conn, pid):
                    self._err(403, "该年度项目已归档，禁止修改"); return
                payload = clean_payload(self._read_body(), "project")
                if not payload:
                    self._err(400, "no fields"); return
                self._ok(services.update(conn, "project", pid, payload))
            elif method == "DELETE" and len(parts) == 1:
                pid = int(parts[0])
                if self._is_archived_project(conn, pid):
                    self._err(403, "该年度项目已归档，禁止删除"); return
                services.soft_delete(conn, "project", pid, self._read_body().get("reason"))
                self._ok({"deleted": pid})
            elif method == "POST" and len(parts) == 2 and parts[1] == "restore":
                self._ok(services.restore(conn, "project", int(parts[0]), self._read_body().get("reason")))
            else:
                self._err(405, "method not allowed")
        except DomainError as exc:
            self._err(403 if '已归档' in str(exc) else 409 if '父项目' in str(exc) else 400, str(exc))
        except ValueError:
            self._err(400, "invalid id")
        finally:
            conn.close()

    # ---------- 资金 ----------
    def _api_funding(self, method, parts, qs):
        self._api_child("funding", method, parts, qs)

    # ---------- 节点 ----------
    def _api_node(self, method, parts, qs):
        self._api_child("node", method, parts, qs)

    def _api_child(self, table, method, parts, qs):
        conn = get_db()
        try:
            if method == "GET" and not parts:
                pid = qs.get("project_id", [None])[0]
                if pid:
                    rows = conn.execute(f"SELECT * FROM {table} WHERE project_id=? AND is_deleted=0 ORDER BY id", (int(pid),)).fetchall()
                else:
                    rows = conn.execute(f"SELECT * FROM {table} WHERE is_deleted=0 ORDER BY id").fetchall()
                self._ok([clean_row(r) for r in rows])
            elif method == "POST" and not parts:
                payload = clean_payload(self._read_body(), table)
                if not payload.get("project_id"):
                    self._err(400, "project_id is required"); return
                if self._is_archived_project(conn, payload["project_id"]):
                    self._err(403, "该项目已归档，禁止新增"); return
                self._ok(services.create(conn, table, payload))
            elif method == "PUT" and len(parts) == 1:
                rid = int(parts[0])
                row = conn.execute(f"SELECT project_id FROM {table} WHERE id=?", (rid,)).fetchone()
                if row and self._is_archived_project(conn, row["project_id"]):
                    self._err(403, "该项目已归档，禁止修改"); return
                payload = clean_payload(self._read_body(), table)
                if not payload:
                    self._err(400, "no fields"); return
                self._ok(services.update(conn, table, rid, payload))
            elif method == "DELETE" and len(parts) == 1:
                rid = int(parts[0])
                row = conn.execute(f"SELECT project_id FROM {table} WHERE id=?", (rid,)).fetchone()
                if row and self._is_archived_project(conn, row["project_id"]):
                    self._err(403, "该项目已归档，禁止删除"); return
                services.soft_delete(conn, table, rid, self._read_body().get("reason"))
                self._ok({"deleted": rid})
            elif method == "POST" and len(parts) == 2 and parts[1] == "restore":
                self._ok(services.restore(conn, table, int(parts[0]), self._read_body().get("reason")))
            else:
                self._err(405, "method not allowed")
        except DomainError as exc:
            self._err(403 if '已归档' in str(exc) else 409 if '父项目' in str(exc) else 400, str(exc))
        except ValueError:
            self._err(400, "invalid id")
        finally:
            conn.close()

    # ---------- 日志 ----------
    def log_message(self, fmt, *args):
        print(f"[{self.command}] {self.path} " + (fmt % args))


def main(open_browser=True):
    init_db()
    try:
        import backup
        backup.auto_backup_if_needed()
    except Exception:
        pass
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    url = f"http://{HOST}:{PORT}"
    print(f"科技项目台账已启动：{url}")
    print("按 Ctrl+C 停止")
    # 自动打开浏览器（延迟，避免和启动抢）
    if open_browser:
        threading.Timer(0.5, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止")


if __name__ == "__main__":
    main(open_browser="--resident" not in sys.argv)
